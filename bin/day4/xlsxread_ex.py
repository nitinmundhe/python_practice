import openpyxl
import csv
F = open('dbdump.csv','r')
rdr = csv.reader(F)
wb = openpyxl.Workbook()
ws1 = wb.active
ws2 = wb.create_sheet('data')
#ws1['A1']=100
#ws1.cell(row=2,column=2).value = 200
#L = [10,20,30]
#ws1.append(L)
h =['ip','date','pics','website']
ws2.append(h)
for r in rdr :
    ws2.append(r)
wb.save('dbdump.xlsx')
wb.close()

F = open('graph.csv','r')
rdr = csv.reader(F)
wb = openpyxl.Workbook()
ws1 = wb.active
for r in rdr :
    ws1.append(r)
wb.save('graph.xlsx')
wb.close()

wb = openpyxl.load_workbook('dbdump.xlsx')
ws3 = wb['data']
ws4 = wb.create_sheet('report')
tr = ws3.max_row
tc = ws3.max_column
for r in range(1,tr+1):
    for c in range(1,tc+1):
        ws4.cell(row=r,column=c).value = ws3.cell(row=r,column=c).value
        #print(ws3.cell(row=r,column=c).value)
# ws3(A1)
# ws3(A)
# ws3(A:C)
# ws3(A1:C10)
# ws3(1:10)

wb.save('dbdump.xlsx')
wb.close()

wb = openpyxl.load_workbook('graph.xlsx')
ws5 = wb.active
ws6 = wb.create_sheet('Bar')
ws7 = wb.create_sheet('Area')
from openpyxl.chart import Reference,AreaChart,BarChart
area = AreaChart()
bar = BarChart()
cat = Reference(ws5,min_row = 1,max_row = 10,min_col = 1)
data = Reference(ws5,min_row = 1,max_row = 10,min_col = 2,max_col = 3)
area.set_categories(cat)
bar.set_categories(cat)
area.add_data(data)
bar.add_data(data)
ws6.add_chart(bar,'A1')
ws7.add_chart(area,'A1')
wb.save('graph.xlsx')
wb.close()