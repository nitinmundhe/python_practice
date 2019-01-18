'''

import re
F = open('C:\nitsmagic\log\server_log.txt','r')
for line in F :
    line =  line.deconde()
    m = re.match('\d\d\d\.\d{3}\.[0-9]{3}]').*\d{2}/[a-zA-Z]{3}/\d{4}.*GET\s+/Pics   ....)

#...
#...
#...
#...  absent for first half till 12:30 day 4


    q = f"insert into logdata values('{IP}','{date}','{image}','{website}')"
    cur.execute(q)
con.commit()
cur.execute('select * from logdata')
cur.fetchall()

import csv
F = open('dbdump.csv','w',newline='')
wt = csv.write(F)
for r in result:
    wt.writerow(r)
F.close()


import csv
import openpyxl
F = open('dbdump.csv','r')
rdr = csv.reader(F)
wb = openpyxl.Workbook()
ws1 = wb.active
ws2 = wb.create_sheet('Data')
ws1['A1']=100
ws1.cell(row=2,column=2).value = 200
L = [10,20,30]
ws1.append(L)
h =['ip','date','pics','website']
ws2.append(h)
for r in rdr :
    ws2.append(r)
wb.save('dbdump.xlsx')
wb.close()
'''

import re

line_org = '"123.123.123.123 - - [26/Apr/2000:00:23:48 -0400] "GET /pics/wpaper.gif HTTP/1.0" 200 6248 "http://www.jafsoft.com/asctortf/" "Mozilla/4.05 (Macintosh; I; PPC)";'
line = line_org = '"123.123.123.123 - - [26/Apr/2000:00:23:48 -0400] "GET /pics/wpaper.gif HTTP/1.0" 200 6248 "http://www.jafsoft.com/asctortf/" "Mozilla/4.05 (Macintosh; I; PPC)";'
#m = re.match('\d\d\d.\d{3}\.[0-9]{3}\.[0-9]{3}]',line,flags=0 )
#m = re.match(r'(\d{3}.){3}\d{3}\s-\s-\s\[\d{1,2}',line,flags=0 )
m = re.match(r'"((\d{3}.){3}\d{3})\s-\s-\s\[\d{1,2}/\w{3}/\d{4}:\d{2}:\d{2}:\d{2}\s+-\d*]\s*"\w{3}\s/pics/\w*.\w*\s*\w*/\d.\d"\s\d*\s\d*',line_org,flags=0 )
if m:
   print(m.group())
