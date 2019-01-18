import csv
import openpyxl
import sqlite3
conn = sqlite3.connect('test.db')
F = open('dbdump.csv','r',newline = '')
rdr = csv.reader(F)
print("Opened database successfully");
for r in rdr :
    print(r)



F = open('dbdump.csv','a',newline = '')
rdr = csv.reader(F)
wb = openpyxl.Workbook()
wb.save('dbdump.xlsx')
ws1 = wb.active
ws2 = wb.create_sheet('Data')
ws1['A1']=100
ws1.cell(row=2,column=2).value = 200
L = [10,20,30]
ws1.append(L)
wb.close()